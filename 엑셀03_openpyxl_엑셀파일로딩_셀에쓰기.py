import  openpyxl  as  op  
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill


wb = op.load_workbook("test2.xlsx") 
ws = wb["직원명부"] 
#"B1" Cell에 입력하기
ws.cell(row=1, column=2).value = "입력테스트1"
#"C1" Cell에 입력하기
ws["C1"].value = "입력테스트2"
#테두리 설정
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

ws["B1"].border = thin_border
ws["C1"].border = thin_border


datalist = [2,4,8,16,32,64,128,256] #임의의 숫자 리스트 정의
light_green_fill     = PatternFill(start_color="FF0000", end_color="FFFFFF", fill_type="solid")

i=5  
for  data  in  datalist:
    ws.cell(row = i, column=1).value = data  #A열(Column=1)에 행을 바꾸면서 입력
    ws.cell(row=i, column=1).fill = light_green_fill
    i=i+1  




wb.save("result2.xlsx") #엑셀 파일 저장

